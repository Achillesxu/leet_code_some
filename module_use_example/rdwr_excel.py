#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@license : (C) Copyright 2013-2017, Easy doesnt enter into grown-up life.
@Software: PyCharm
@Project : algo_exe
@Time : 2017/12/29 下午1:56
@Author : achilles_xushy
@contact : yuqingxushiyin@gmail.com
@Site : 
@File : rdwr_excel.py
@desc :
"""
import string
from openpyxl import load_workbook
from random import randint


def func1(in_file, out_file, in_name):
    wb = load_workbook(in_file)
    sheet = wb.active
    row_range = []
    name_list = []
    for i in range(1, sheet.max_row+1):
        for j in range(sheet.max_column):
            if sheet[string.ascii_uppercase[j]+str(i)].data_type == 'n' and \
                    sheet[string.ascii_uppercase[j] + str(i)].value:
                row_range.append(sheet[string.ascii_uppercase[j]+str(i)].row)
                break

    for i in row_range:
        name_list.append(sheet[string.ascii_uppercase[1]+str(i)].value)
    print(name_list)
    # set value
    for i in row_range:
        for j in range(3, sheet.max_column):
            if sheet[string.ascii_uppercase[1] + str(i)].value == in_name:
                break
            sheet[string.ascii_uppercase[j]+str(i)] = randint(95, 99)
    wb.save(out_file)


if __name__ == '__main__':
    xls_file = '/Users/achilles_xushy/Desktop/员工互评-云端.xlsx'
    xls_file_out = '/Users/achilles_xushy/Desktop/员工互评-云端1.xlsx'
    except_name = '徐世寅'
    func1(xls_file, xls_file_out, except_name)

    # data = {
    #     'police_status': '1',
    #     'signature': hashlib.md5(reduce(operator.add,
    #  ['23'.encode(), '1'.encode(), '3bc7e1bf41c5e8a7085377db1af1aef0'.encode()])).hexdigest()
    # }
    # ret = requests.post('http://store.7po.com/police/inspection/interface', data=data)
    # print(ret.status_code)

